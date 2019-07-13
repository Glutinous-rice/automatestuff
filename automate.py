import openpyxl
wb=openpyxl.load_workbook('updatedProduceSales.xlsx')
ws=wb['Sheet']

#需要改动的农产品及其价格
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for rowNum in range(2,ws.max_row+1):   #遍历农产品名字那一列
    produceName=ws.cell(row=rowNum,column=1).value         #把单元格内容复制给produnceName
    if produceName in PRICE_UPDATES:                       #如果单元格的内容在所需要改的目录之中
        ws.cell(row=rowNum,column=2).value=PRICE_UPDATES[produceName]
        #就把需要改的价格赋值给对应的价格栏

wb.save('updateautomate.xlsx')